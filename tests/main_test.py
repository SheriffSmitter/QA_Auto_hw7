import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_read_xlsx_file():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("file_example_XLSX_50.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=3, column=7).value
            name = "16/08/2016"
            assert name in cell_value


def test_csv():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("import_empl_csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[10]
            result_list = second_row[0].split(';')
            vendor = "B4COM"
            model = "CS4148D-6U"
            assert result_list[2] == vendor
            assert result_list[3] == model


def test_read_pdf_file():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open("pythonx.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]
            text = page.extract_text()
            assert 'Rapid' in text
